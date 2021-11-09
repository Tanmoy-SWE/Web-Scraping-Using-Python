import openpyxl as op
wb = op.load_workbook('workersData.xlsx')
#Storingthe sheetNames in a list
sheetNames = wb.sheetnames
print(sheetNames)
adminList = wb['WOs']
print(adminList.max_row)
print(adminList.max_column)
# data = adminList.cell(row=2,column=1).value
# print(data)
# for i in range (1,100):
#     print(adminList.cell(row=i,column=1).value)
firstRow = (list)(adminList.columns)[1]
print(firstRow[0].value)
